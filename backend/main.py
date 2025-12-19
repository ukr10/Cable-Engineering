from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Tuple, Any
import math
from datetime import datetime
import io
import json
import os

# SQLAlchemy for persistence
try:
    from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, JSON, ForeignKey
    from sqlalchemy.orm import sessionmaker, declarative_base
    from sqlalchemy import text
    from sqlalchemy.exc import OperationalError
except Exception:
    create_engine = None
    declarative_base = None

# Excel handling
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except ImportError:
    load_workbook = None

# Graph-based routing
try:
    import networkx as nx
except ImportError:
    nx = None

app = FastAPI(title="SCEAP - Backend API", version="1.0.0")

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== Data Models ====================

class CableInput(BaseModel):
    cable_number: str
    description: str = ''
    load_kw: float = 0.0
    load_kva: float = 0.0
    voltage: float = 415.0
    pf: float = 0.9
    efficiency: float = 0.95
    length: float = 0.0
    runs: int = 1
    cable_type: str = 'C'
    from_equipment: str = ''
    to_equipment: str = ''
    # optional additional columns
    breaker_type: Optional[str] = None
    feeder_type: Optional[str] = None
    cores: Optional[int] = 3
    quantity: Optional[int] = 1
    voltage_variation: Optional[float] = None
    power_supply: Optional[str] = None
    installation: Optional[str] = None
    # Advanced fields
    prospective_sc: Optional[float] = None  # prospective short-circuit current (A)
    phase_type: Optional[str] = None  # 'single' or 'three'
    ambient_temp: Optional[float] = None  # in °C

class CableResult(BaseModel):
    id: str
    cable_number: str
    description: str
    flc: float
    derated_current: float
    selected_size: float
    voltage_drop: float
    sc_check: bool
    grouping_factor: float
    status: str = "pending"
    cores: int = 3
    od: float = 0.0
    breaker_type: Optional[str] = None
    feeder_type: Optional[str] = None
    quantity: Optional[int] = None
    # additional diagnostics
    ampacity: Optional[float] = None
    ampacity_base: Optional[float] = None
    ampacity_corrected: Optional[float] = None
    ampacity_margin: Optional[float] = None
    ampacity_margin_pct: Optional[float] = None
    vd_limit: Optional[float] = None
    vd_pass: Optional[bool] = None
    ao: Optional[bool] = None
    resistance_per_m: Optional[float] = None
    prospective_sc: Optional[float] = None
    standard: Optional[str] = None
    recommended_cores: Optional[int] = None
    recommended_runs: Optional[int] = None
    resistance_per_m: Optional[float] = None
    reactance_per_m: Optional[float] = None
    configuration: Optional[str] = None
    standard_ref: Optional[str] = None
    formulas: Optional[Dict[str, str]] = None

class ExcelUploadResponse(BaseModel):
    success: bool
    cables_imported: int
    errors: List[str] = []
    results: List[CableResult] = []
    inputs: Optional[List[Dict[str, Any]]] = None
from fastapi.responses import StreamingResponse

class RoutingRequest(BaseModel):
    cable_id: str
    from_equipment: str
    to_equipment: str
    cable_od: float
    cable_csa: float
    runs: int

class RoutingResult(BaseModel):
    cable_id: str
    path: List[str]
    total_length: float
    fill_status: Dict[str, float]
    warnings: List[str] = []

class ProjectSetup(BaseModel):
    project_name: str
    plant_type: str
    standard: str
    voltage_levels: List[float]
    service_condition: str

# ==================== Cable Sizing Calculations ====================

def calculate_flc(load_kw: float, voltage: float, pf: float, efficiency: float) -> float:
    """Calculate Full Load Current: I = P / (√3 × V × PF × η)"""
    if voltage == 0 or pf == 0:
        return 0
    # load_kw is in kW, convert to Watts
    return (load_kw * 1000.0) / (math.sqrt(3) * voltage * pf * efficiency)


def calculate_flc_for_phase(load_kw: float, voltage: float, pf: float, efficiency: float, phase: str = 'three') -> float:
    """Calculate FLC taking into account single-phase or three-phase system."""
    if phase == 'single' or phase == '1' or phase == '1p':
        if voltage == 0 or pf == 0:
            return 0
        return (load_kw * 1000.0) / (voltage * pf * efficiency)
    # default to three-phase
    return calculate_flc(load_kw, voltage, pf, efficiency)

def apply_derating(current: float, grouping_factor: float = 0.8, temp_factor: float = 0.95, 
                   installation_factor: float = 1.0) -> float:
    """Apply derating factors"""
    return current / (grouping_factor * temp_factor * installation_factor)


def derive_grouping_factor(runs: int, feeder_type: Optional[str] = None, phase: str = 'three') -> float:
    """Derive a grouping factor from runs and feeder type. Simplified heuristics."""
    gf = 0.8
    if phase == 'single' or phase.startswith('1'):
        gf = 0.9
    if runs <= 1:
        gf *= 1.0
    elif runs == 2:
        gf *= 0.9
    elif runs == 3:
        gf *= 0.85
    else:
        gf *= 0.8

    if feeder_type and (str(feeder_type).lower().startswith('f') or str(feeder_type).lower().startswith('feeder')):
        gf *= 0.95
    return max(0.5, round(gf, 2))


def derive_temp_factor(ambient_temp: Optional[float]) -> float:
    """Approximate derating factor based on ambient temperature (per 10°C +5% derating approx.)"""
    if ambient_temp is None:
        return 0.95
    base = 30.0
    delta = ambient_temp - base
    steps = int(delta / 10.0)
    f = 1.0
    for _ in range(max(0, steps)):
        f *= 0.95
    return round(f, 3)


def adiabatic_short_circuit_capacity(size_mm2: float, t: float = 1.0, k: float = 115.0) -> float:
    """Adiabatic short circuit capacity approx for copper: I = K * sqrt(S/t) (I in A)
    K typical ~ 115 (copper), t is clearing time in seconds
    """
    try:
        return k * math.sqrt(size_mm2 / (t if t > 0 else 1.0))
    except Exception:
        return 0.0

def calculate_voltage_drop(flc: float, length: float, voltage: float, runs: int = 1, 
                          resistance: Optional[float] = None, size_mm2: Optional[float] = None, reactance: Optional[float] = None) -> float:
    """Calculate voltage drop percentage.

    Uses magnitude of impedance (R and X) when both are available:
    Vd% = (√3 × I_per_run × L × √(R^2 + X^2)) / V × 100
    Falls back to resistance-only calculation when reactance is None.
    """
    if voltage == 0:
        return 0
    # If resistance is not supplied, derive from conductor size
    if resistance is None and size_mm2 is not None:
        resistance_per_m = get_resistance_per_m(size_mm2)
    elif resistance is not None:
        resistance_per_m = resistance
    else:
        resistance_per_m = 0.02

    # With multiple runs (parallel conductors), current per conductor decreases
    i_per_run = flc / (runs if runs and runs > 0 else 1)
    if reactance is not None:
        z = math.sqrt((resistance_per_m or 0.0) ** 2 + (reactance) ** 2)
        vd = (math.sqrt(3) * i_per_run * length * z) / voltage
    else:
        vd = (math.sqrt(3) * i_per_run * length * resistance_per_m) / voltage
    return vd * 100

def select_cable_size(derated_current: float, standard: str = "IEC", catalog_name: Optional[str] = None, grouping: float = 1.0, temp_factor: float = 1.0) -> Tuple[str, float, float, Optional[float], Optional[int], Optional[int]]:
    """Select cable size and return (size_string, size_mm2)"""
    # If a catalog_name is provided, load catalog from DB/in-memory
    catalog_list = None
    if catalog_name:
        if DB_USABLE:
            db = SessionLocal()
            try:
                try:
                    row = db.query(CatalogModel).filter(CatalogModel.name == catalog_name).first()
                    if row and row.payload:
                        catalog_list = row.payload
                except OperationalError:
                    # Table missing or DB inaccessible: fallback to in-memory catalog
                    catalog_list = CATALOG_STORE.get(catalog_name)
            finally:
                db.close()
        else:
            catalog_list = CATALOG_STORE.get(catalog_name)

    if not catalog_list:
        # fallback to default IEC table
        cable_sizes_iec = [
        (10, 1.5), (16, 2.5), (25, 4), (35, 6), (50, 10),
        (63, 16), (80, 25), (100, 35), (125, 50), (160, 70),
        (200, 95), (250, 120), (315, 150), (400, 185), (500, 240)
    ]
    
        for amps, size in cable_sizes_iec:
            if derated_current <= amps:
                res = get_resistance_per_m(size)
                return (f"3 x {size}", size, amps, res, 1, 3)
        # If derated_current exceeds table, compute parallel runs of largest conductor
        max_amps, max_size = cable_sizes_iec[-1]
        runs_needed = int(math.ceil(derated_current / max_amps)) if max_amps > 0 else None
        res = get_resistance_per_m(max_size)
        return (f"3 x {max_size} (runs={runs_needed})", max_size, float(max_amps), res, runs_needed, 3)
    else:
        # If catalog_list exists, try to find appropriate size by ampacity
        if catalog_list:
            # Try to find a single conductor that meets derated_current, else compute runs
            best_option = None
            # generate options for multicore and single-core paralleled approaches
            best_option = None
            for entry in catalog_list:
                size = entry.get('size_mm2') or entry.get('size')
                if not size:
                    continue
                base_amp, amp_corrected, _src = compute_ampacity_from_entry(entry, standard=standard, grouping_factor=grouping, temp_factor=temp_factor)
                amp = amp_corrected or base_amp or 0
                # Option: single run of this cable
                if amp and amp >= derated_current:
                    res = entry.get('resistance_per_m')
                    return (f"{int(entry.get('cores') or 3)}C x {size} mm²", float(size), float(base_amp or amp), res, 1, int(entry.get('cores') or 3))

                # Option: compute runs required (paralleling)
                if amp and amp > 0:
                    runs = int(math.ceil(derated_current / amp))
                else:
                    runs = None

                par_allowed = entry.get('paralleled_count') or None
                # if par_allowed present and runs > par_allowed, mark as not allowed
                if par_allowed and runs and runs > par_allowed:
                    allowed = False
                else:
                    allowed = True

                option = {'entry': entry, 'runs': runs, 'size': float(size), 'amp': float(amp or 0), 'allowed': allowed}
                if not allowed:
                    continue
                if best_option is None:
                    best_option = option
                else:
                    # prefer fewer runs, then smaller size
                    if option['runs'] is not None and best_option['runs'] is not None:
                        if option['runs'] < best_option['runs'] or (option['runs'] == best_option['runs'] and option['size'] < best_option['size']):
                            best_option = option
                    elif option['runs'] is not None:
                        best_option = option
            if best_option:
                entry = best_option['entry']
                res = entry.get('resistance_per_m')
                return (f"{int(entry.get('cores') or 3)}C x {best_option['size']} mm² (runs={best_option['runs']})", best_option['size'], best_option['amp'], res, int(best_option['runs']) if best_option['runs'] is not None else None, int(entry.get('cores') or 3))
    
    return ("3 x 300", 300, 0, None, None, 3)

def check_short_circuit(cable_size: float) -> bool:
    """Simplified short circuit check"""
    return cable_size >= 1.5

def calculate_cable_od(cores: int, csa: float) -> float:
    """Estimate cable outer diameter based on cores and CSA"""
    base = math.sqrt(cores * csa) * 1.5
    return round(base, 1)


def get_resistance_per_m(size_mm2: float) -> float:
    """Approximate resistance per meter (Ohm/m) for copper conductors at 20°C.
    Values are based on typical tabulated resistances (ohm/km) / 1000.
    """
    table_ohm_per_km = {
        1.5: 12.1,
        2.5: 7.41,
        4: 4.61,
        6: 3.08,
        10: 1.83,
        16: 1.15,
        25: 0.727,
        35: 0.524,
        50: 0.387,
        70: 0.268,
        95: 0.193,
        120: 0.153,
        150: 0.124,
        185: 0.101,
        240: 0.075
    }
    # Find nearest key
    if size_mm2 in table_ohm_per_km:
        return table_ohm_per_km[size_mm2] / 1000.0
    # fallback: choose nearest smaller or default to 0.02 ohm/m
    keys = sorted(table_ohm_per_km.keys())
    for k in keys:
        if size_mm2 <= k:
            return table_ohm_per_km[k] / 1000.0
    return table_ohm_per_km[keys[-1]] / 1000.0


def compute_ampacity_from_entry(entry: Dict[str, Any], standard: str = 'IEC', grouping_factor: float = 1.0, temp_factor: float = 1.0, installation_factor: float = 1.0) -> Tuple[Optional[float], Optional[float], str]:
    """Compute corrected ampacity for a catalog `entry` using derating factors.

    Returns (corrected_ampacity, source_description).
    """
    base_amp = None
    source = 'none'
    if entry is None:
        return None, source

    # Prefer air ampacity if present, else ground, else None
    if entry.get('ampacity_air'):
        base_amp = float(entry.get('ampacity_air'))
        source = 'ampacity_air'
    elif entry.get('ampacity_ground'):
        base_amp = float(entry.get('ampacity_ground'))
        source = 'ampacity_ground'
    elif entry.get('ampacity'):
        base_amp = float(entry.get('ampacity'))
        source = 'ampacity'
    elif entry.get('size_mm2'):
        # fallback: map size to IEC table default amps
        size = float(entry.get('size_mm2'))
        # simple default mapping from size to typical amps
        default_map = {1.5:10,2.5:16,4:25,6:35,10:50,16:63,25:80,35:100,50:125,70:160,95:200,120:250,150:315,185:400,240:500}
        # find nearest >= size
        chosen = None
        keys = sorted(default_map.keys())
        for k in keys:
            if size <= k:
                chosen = default_map[k]
                break
        if not chosen:
            chosen = default_map[keys[-1]]
        base_amp = float(chosen)
        source = 'default_table'

    if base_amp is None:
        return None, None, source

    corrected = base_amp * grouping_factor * temp_factor * installation_factor
    if standard and standard.lower().startswith('is'):
        # IS often uses slightly stricter corrections
        corrected = corrected * 0.95
        source += '->IS_adjusted'

    return base_amp, corrected, source

# ==================== Excel Import ====================

def parse_excel_cables(file_content: bytes) -> Tuple[List[CableInput], List[str]]:
    """Parse Excel file and extract cable data"""
    errors = []
    cables = []
    
    if not load_workbook:
        errors.append("openpyxl not installed")
        return cables, errors
    
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        # Build header map from first row
        header_row = [str(cell).strip() if cell is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = [h.lower().strip().replace(' ', '_').replace('.', '').replace('-', '_') for h in header_row]
        idx_map = {name: i for i, name in enumerate(headers)}

        def get_val(row, key):
            i = idx_map.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        # Basic header validation
        if not any(k in idx_map for k in ('cable_number', 'cable_no', 'cable')):
            errors.append('Missing key column: cable_number/cable_no/cable')
            return [], errors
        if not any(k in idx_map for k in ('load_kw', 'kw', 'load')):
            # not fatal, add warning but allow import
            errors.append('Warning: load_kw/kw/load column missing or unmapped; assuming 0 kW for rows.')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:
                    continue
                
                cable = CableInput(
                    cable_number=str(get_val(row, 'cable_number') or get_val(row, 'cable_no') or get_val(row, 'cable')),
                    description=str(get_val(row, 'description') or get_val(row, 'desc') or '') if (get_val(row, 'description') or get_val(row, 'desc')) else '',
                    load_kw=float(get_val(row, 'load_kw') or get_val(row, 'kw') or get_val(row, 'kW') or get_val(row, 'load')) if (get_val(row, 'load_kw') or get_val(row, 'kw') or get_val(row, 'kW') or get_val(row, 'load')) else 0,
                    load_kva=float(get_val(row, 'load_kva') or get_val(row, 'kva') or 0) if (get_val(row, 'load_kva') or get_val(row, 'kva')) else 0,
                    voltage=float(get_val(row, 'voltage') or get_val(row, 'v') or 415) if (get_val(row, 'voltage') or get_val(row, 'v')) else 415,
                    pf=float(get_val(row, 'pf') or 0.9) if (get_val(row, 'pf')) else 0.9,
                    efficiency=float(get_val(row, 'efficiency') or 0.95) if (get_val(row, 'efficiency')) else 0.95,
                    length=float(get_val(row, 'length') or 0) if (get_val(row, 'length')) else 0,
                    runs=int(get_val(row, 'runs') or 1) if (get_val(row, 'runs')) else 1,
                    cable_type=str(get_val(row, 'cable_type') or 'C') if (get_val(row, 'cable_type')) else 'C',
                    from_equipment=str(get_val(row, 'from_equipment') or get_val(row, 'from') or '') if (get_val(row, 'from_equipment') or get_val(row, 'from')) else '',
                    to_equipment=str(get_val(row, 'to_equipment') or get_val(row, 'to') or '') if (get_val(row, 'to_equipment') or get_val(row, 'to')) else '',
                    breaker_type=str(get_val(row, 'breaker_type') or get_val(row, 'breaker') or '') if (get_val(row, 'breaker_type') or get_val(row, 'breaker')) else None,
                    feeder_type=str(get_val(row, 'feeder_type') or '') if get_val(row, 'feeder_type') else None,
                    cores=int(get_val(row, 'cores') or 3) if get_val(row, 'cores') else 3,
                    quantity=int(get_val(row, 'quantity') or 1) if get_val(row, 'quantity') else 1,
                    voltage_variation=float(get_val(row, 'voltage_variation') or 0) if get_val(row, 'voltage_variation') else None,
                    power_supply=str(get_val(row, 'power_supply') or '') if get_val(row, 'power_supply') else None,
                    installation=str(get_val(row, 'installation') or '') if get_val(row, 'installation') else None,
                    prospective_sc=float(get_val(row, 'prospective_sc') or get_val(row, 'prospective_sc_ka') or 0) if (get_val(row, 'prospective_sc') or get_val(row, 'prospective_sc_ka')) else None,
                    phase_type=str(get_val(row, 'phase') or get_val(row, 'phase_type') or 'three') if (get_val(row, 'phase') or get_val(row, 'phase_type')) else None,
                    ambient_temp=float(get_val(row, 'ambient_temp') or get_val(row, 'ambient_temperature') or 0) if (get_val(row, 'ambient_temp') or get_val(row, 'ambient_temperature')) else None,
                )
                cables.append(cable)
            except (ValueError, IndexError, TypeError) as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return cables, errors
    except Exception as e:
        errors.append(f"Error: {str(e)}")
        return cables, errors


def parse_catalog_excel(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Parse catalog excel into a list of dicts.

    Expected/accepted headers (case-insensitive examples):
      - size_mm2 or size
      - cores (e.g., 3, 3.5, 4)
      - ampacity_air or air
      - ampacity_ground or ground
      - resistance_90c_ohm_per_km or resistance_ohm_per_km
      - reactance_ohm_per_km or reactance
      - cable_dia_mm or cable_dia
      - xlpe
      - grouping_k2

    This parser normalizes values and returns entries with ohm/m units for resistance/reactance.
    """
    errors = []
    catalog = []
    if not load_workbook:
        errors.append("openpyxl not installed")
        return catalog, errors
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        # Build header map from first row
        header_row = [str(cell).strip() if cell is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = [h.lower().strip().replace(' ', '_').replace('.', '').replace('-', '_') for h in header_row]
        idx_map = {name: i for i, name in enumerate(headers)}

        def get_cell(row, key):
            i = idx_map.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(cell is not None for cell in row):
                    continue

                size = get_cell(row, 'size_mm2') or get_cell(row, 'size')
                cores = get_cell(row, 'cores') or get_cell(row, 'no_of_cores')
                amp_air = get_cell(row, 'ampacity_air') or get_cell(row, 'air') or get_cell(row, 'current_rating_of_the_cable')
                amp_ground = get_cell(row, 'ampacity_ground') or get_cell(row, 'ground')
                res_km = get_cell(row, 'resistance_90deg_ohm_per_km') or get_cell(row, 'resistance_90deg_ohmperkm') or get_cell(row, 'resistance_ohm_per_km') or get_cell(row, 'resistance')
                react_km = get_cell(row, 'reactance_ohm_per_km') or get_cell(row, 'reactance')
                dia = get_cell(row, 'cable_dia_mm') or get_cell(row, 'cable_dia')
                xlpe = get_cell(row, 'xlpe')
                pairs = get_cell(row, 'pairs') or get_cell(row, 'no_of_pairs')
                paralleled = get_cell(row, 'paralleled') or get_cell(row, 'paralleled_count') or get_cell(row, 'parallels')
                conductor_material = get_cell(row, 'conductor_material') or get_cell(row, 'material')
                insulation = get_cell(row, 'insulation') or xlpe
                sheath = get_cell(row, 'sheath') or get_cell(row, 'armour')
                formation = get_cell(row, 'formation')
                grouping_k2 = get_cell(row, 'grouping_k2')

                # Normalize and cast
                try:
                    size_val = float(size) if size is not None and str(size).strip() != '' else None
                except Exception:
                    size_val = None
                try:
                    cores_val = float(cores) if cores is not None and str(cores).strip() != '' else None
                except Exception:
                    cores_val = None
                try:
                    amp_air_val = float(amp_air) if amp_air is not None and str(amp_air).strip() != '' else None
                except Exception:
                    amp_air_val = None
                try:
                    amp_ground_val = float(amp_ground) if amp_ground is not None and str(amp_ground).strip() != '' else None
                except Exception:
                    amp_ground_val = None
                try:
                    res_m = float(res_km) / 1000.0 if res_km is not None and str(res_km).strip() != '' else None
                except Exception:
                    res_m = None
                try:
                    react_m = float(react_km) / 1000.0 if react_km is not None and str(react_km).strip() != '' else None
                except Exception:
                    react_m = None
                try:
                    dia_val = float(dia) if dia is not None and str(dia).strip() != '' else None
                except Exception:
                    dia_val = None

                entry = {
                    'size_mm2': size_val,
                    'cores': int(cores_val) if cores_val is not None else None,
                    'ampacity_air': amp_air_val,
                    'ampacity_ground': amp_ground_val,
                    'resistance_per_m': res_m,
                    'reactance_per_m': react_m,
                    'cable_dia_mm': dia_val,
                    'xlpe': str(xlpe).strip() if xlpe is not None else None,
                    'grouping_k2': float(grouping_k2) if grouping_k2 is not None and str(grouping_k2).strip() != '' else None,
                    'pairs': int(pairs) if pairs is not None and str(pairs).strip() != '' else None,
                    'paralleled_count': int(paralleled) if paralleled is not None and str(paralleled).strip() != '' else None,
                    'conductor_material': str(conductor_material).strip() if conductor_material is not None else None,
                    'insulation': str(insulation).strip() if insulation is not None else None,
                    'sheath': str(sheath).strip() if sheath is not None else None,
                    'formation': str(formation).strip() if formation is not None else None,
                }
                catalog.append(entry)
            except (ValueError, IndexError, TypeError) as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        # Sort by size then ampacity
        catalog = sorted(catalog, key=lambda x: (x.get('size_mm2') or 0, x.get('ampacity_air') or 0))
        return catalog, errors
    except Exception as e:
        errors.append(f"Error: {str(e)}")
        return catalog, errors

# ==================== Routing Engine ====================

class TrayNetwork:
    """Tray network for routing calculations"""
    
    def __init__(self):
        self.graph = nx.Graph() if nx else None
        self.tray_data = {}
        self._build_sample_network()
    
    def _build_sample_network(self):
        """Build sample tray network"""
        if not self.graph:
            return
        
        trays = ["PHF-01", "PHF-02", "PHF-03", "PHF-04", "PHF-05", "DB-01", "DB-02"]
        equipment = ["Transformer", "Panel A", "Panel B", "Motor M1", "Motor M2"]
        
        all_nodes = trays + equipment
        self.graph.add_nodes_from(all_nodes)
        
        connections = [
            ("Transformer", "PHF-01", 10),
            ("PHF-01", "PHF-02", 5),
            ("PHF-02", "PHF-03", 8),
            ("PHF-03", "DB-01", 6),
            ("DB-01", "Panel A", 12),
            ("DB-01", "Panel B", 15),
            ("PHF-02", "PHF-04", 7),
            ("PHF-04", "DB-02", 9),
            ("DB-02", "Motor M1", 10),
            ("DB-02", "Motor M2", 11),
        ]
        
        for src, dst, dist in connections:
            self.graph.add_edge(src, dst, weight=dist)
        
        self.tray_data = {
            "PHF-01": {"fill": 45, "capacity": 1000},
            "PHF-02": {"fill": 60, "capacity": 1000},
            "PHF-03": {"fill": 70, "capacity": 1000},
            "PHF-04": {"fill": 30, "capacity": 1000},
            "DB-01": {"fill": 55, "capacity": 800},
            "DB-02": {"fill": 65, "capacity": 800},
        }
    
    def find_shortest_path(self, source: str, target: str) -> Tuple[List[str], float]:
        """Find shortest path"""
        if not self.graph:
            return [source, target], 50.0
        
        try:
            path = nx.shortest_path(self.graph, source, target, weight='weight')
            length = nx.shortest_path_length(self.graph, source, target, weight='weight')
            return path, float(length)
        except:
            return [source, target], 50.0
    
    def find_least_fill_path(self, source: str, target: str) -> Tuple[List[str], float]:
        """Find least-fill path: penalize edges that connect to high-fill trays."""
        if not self.graph:
            return [source, target], 50.0

        # Build a temporary weighted graph where edge weight = base_weight + fill_penalty
        G = self.graph.copy()

        for u, v, data in G.edges(data=True):
            base = float(data.get('weight', 1.0))
            fill_u = float(self.tray_data.get(u, {}).get('fill', 0))
            fill_v = float(self.tray_data.get(v, {}).get('fill', 0))
            # penalty scale: each percent fill adds 0.2 units
            penalty = ((fill_u + fill_v) / 2.0) * 0.2
            data['weight'] = base + penalty

        try:
            path = nx.shortest_path(G, source, target, weight='weight')
            length = nx.shortest_path_length(G, source, target, weight='weight')
            return path, float(length)
        except Exception:
            return [source, target], 50.0

routing_engine = TrayNetwork()

# ==================== Database / Persistence ====================

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///./sceap.db')
print(f"[startup] Using DATABASE_URL={DATABASE_URL}")
# If the default sqlite file exists but is not writable by this process, switch to a local fallback
if DATABASE_URL.startswith('sqlite:///'):
    dbfile = DATABASE_URL.replace('sqlite:///', '')
    try:
        if os.path.exists(dbfile) and not os.access(dbfile, os.W_OK):
            # create a per-user fallback DB in workspace
            fallback = './sceap_dev.db'
            print(f"[startup] WARNING: DB file {dbfile} not writable. Switching DB to {fallback}")
            DATABASE_URL = f"sqlite:///{fallback}"
    except Exception:
        pass
DB_AVAILABLE = create_engine is not None
DB_USABLE = False
Base = declarative_base() if declarative_base else None
Engine = None
SessionLocal = None

if DB_AVAILABLE:
    Engine = create_engine(DATABASE_URL, echo=False, future=True)
    SessionLocal = sessionmaker(bind=Engine, autoflush=False, autocommit=False)

    class ProjectModel(Base):
        __tablename__ = 'projects'
        id = Column(Integer, primary_key=True, index=True)
        project_id = Column(String, unique=True, index=True)
        name = Column(String)
        plant_type = Column(String)
        standard = Column(String)
        voltage_levels = Column(JSON)
        service_condition = Column(String)
        created_at = Column(DateTime)

    class CableResultModel(Base):
        __tablename__ = 'cable_results'
        id = Column(Integer, primary_key=True, index=True)
        result_id = Column(String, index=True)
        project_id = Column(String, ForeignKey('projects.project_id'), index=True)
        cable_number = Column(String)
        payload = Column(JSON)
        created_at = Column(DateTime)
    class CatalogModel(Base):
        __tablename__ = 'catalogs'
        id = Column(Integer, primary_key=True, index=True)
        name = Column(String, unique=True, index=True)
        payload = Column(JSON)
        created_at = Column(DateTime)

    def init_db():
            try:
                Base.metadata.create_all(bind=Engine)
            except Exception:
                pass

    def is_db_writable() -> bool:
        """Attempt to write a minimal transaction to determine DB usability."""
        global DB_USABLE
        if not DB_AVAILABLE or Engine is None:
            DB_USABLE = False
            return False
        try:
            # Attempt to create tables and a session
            Base.metadata.create_all(bind=Engine)
            db = SessionLocal()
            try:
                # Try a lightweight query
                db.execute(text("SELECT 1"))
                db.rollback()
            finally:
                db.close()
            DB_USABLE = True
            return True
        except Exception:
            import traceback
            print("[startup] DB usability check failed:")
            traceback.print_exc()
            DB_USABLE = False
            return False

else:
    def init_db():
        return

# In-memory fallback for catalogs when DB is not available
CATALOG_STORE: Dict[str, List[Dict[str, Any]]] = {}
INMEM_PROJECTS: Dict[str, Dict[str, Any]] = {}
INMEM_CABLE_RESULTS: Dict[str, Dict[str, Any]] = {}

# Default catalog (IEC-like)
DEFAULT_CATALOG = [
    { 'size_mm2': 1.5, 'cores': 3, 'formation': '3C', 'ampacity_air': 10, 'ampacity_ground': 12, 'resistance_per_m': 0.0121, 'reactance_per_m': 0.00011, 'cable_dia_mm': 4.5 },
    { 'size_mm2': 2.5, 'cores': 3, 'formation': '3C', 'ampacity_air': 16, 'ampacity_ground': 18, 'resistance_per_m': 0.00741, 'reactance_per_m': 0.00009, 'cable_dia_mm': 5.0 },
    { 'size_mm2': 4, 'cores': 3, 'formation': '3C', 'ampacity_air': 25, 'ampacity_ground': 27, 'resistance_per_m': 0.00461, 'reactance_per_m': 0.00008, 'cable_dia_mm': 6.0 },
    { 'size_mm2': 6, 'cores': 3, 'formation': '3C', 'ampacity_air': 35, 'ampacity_ground': 37, 'resistance_per_m': 0.00308, 'reactance_per_m': 0.000075, 'cable_dia_mm': 7.0 },
    { 'size_mm2': 10, 'cores': 3, 'formation': '3C', 'ampacity_air': 50, 'ampacity_ground': 55, 'resistance_per_m': 0.00183, 'reactance_per_m': 0.000065, 'cable_dia_mm': 9.0 },
    { 'size_mm2': 16, 'cores': 3, 'formation': '3C', 'ampacity_air': 63, 'ampacity_ground': 67, 'resistance_per_m': 0.00115, 'reactance_per_m': 0.000058, 'cable_dia_mm': 10.5 },
    { 'size_mm2': 25, 'cores': 3, 'formation': '3C', 'ampacity_air': 80, 'ampacity_ground': 85, 'resistance_per_m': 0.000727, 'reactance_per_m': 0.000052, 'cable_dia_mm': 12.0 },
    { 'size_mm2': 35, 'cores': 3, 'formation': '3C', 'ampacity_air': 100, 'ampacity_ground': 104, 'resistance_per_m': 0.000524, 'reactance_per_m': 0.000047, 'cable_dia_mm': 14.0 },
    { 'size_mm2': 50, 'cores': 3, 'formation': '3C', 'ampacity_air': 125, 'ampacity_ground': 130, 'resistance_per_m': 0.000387, 'reactance_per_m': 0.000042, 'cable_dia_mm': 16.0 },
    { 'size_mm2': 70, 'cores': 3, 'formation': '3C', 'ampacity_air': 160, 'ampacity_ground': 165, 'resistance_per_m': 0.000268, 'reactance_per_m': 0.000038, 'cable_dia_mm': 18.0 },
    { 'size_mm2': 95, 'cores': 3, 'formation': '3C', 'ampacity_air': 200, 'ampacity_ground': 205, 'resistance_per_m': 0.000193, 'reactance_per_m': 0.000035, 'cable_dia_mm': 21.0 },
    { 'size_mm2': 120, 'cores': 3, 'formation': '3C', 'ampacity_air': 250, 'ampacity_ground': 255, 'resistance_per_m': 0.000153, 'reactance_per_m': 0.000033, 'cable_dia_mm': 23.0 },
    { 'size_mm2': 150, 'cores': 3, 'formation': '3C', 'ampacity_air': 315, 'ampacity_ground': 320, 'resistance_per_m': 0.000124, 'reactance_per_m': 0.00003, 'cable_dia_mm': 27.0 },
    { 'size_mm2': 185, 'cores': 3, 'formation': '3C', 'ampacity_air': 350, 'ampacity_ground': 355, 'resistance_per_m': 0.000101, 'reactance_per_m': 0.000028, 'cable_dia_mm': 30.0 },
    { 'size_mm2': 240, 'cores': 3, 'formation': '3C', 'ampacity_air': 400, 'ampacity_ground': 405, 'resistance_per_m': 0.000075, 'reactance_per_m': 0.000025, 'cable_dia_mm': 34.0 },
    # Some single-core entries often used for paralleled runs
    { 'size_mm2': 185, 'cores': 1, 'formation': '1C', 'ampacity_air': 271, 'ampacity_ground': 305, 'resistance_per_m': 0.00021, 'reactance_per_m': 0.000072, 'cable_dia_mm': 36.0 },
    { 'size_mm2': 240, 'cores': 1, 'formation': '1C', 'ampacity_air': 448, 'ampacity_ground': 385, 'resistance_per_m': 0.00018, 'reactance_per_m': 0.0000719, 'cable_dia_mm': 44.0 },
]



@app.on_event('startup')
def on_startup():
    init_db()
    # Check DB write usability; if not usable we will fall back to in-memory stores
    is_db_writable()
    # Re-run create_all to ensure any new models are created (handles reloads)
    try:
        if DB_AVAILABLE and Engine is not None and Base is not None:
            Base.metadata.create_all(bind=Engine)
    except Exception:
        # If DB operations failed, ensure an in-memory fallback default catalog exists
        if 'default' not in CATALOG_STORE:
            CATALOG_STORE['default'] = DEFAULT_CATALOG
        pass
    # ensure default catalog exists
    try:
        if DB_USABLE:
            db = SessionLocal()
            try:
                # If the catalogs table doesn't exist yet, try to create it before querying
                try:
                    existing = db.query(CatalogModel).filter(CatalogModel.name == 'default').first()
                except OperationalError:
                    # Attempt a create_all to add missing tables then retry once
                    try:
                        Base.metadata.create_all(bind=Engine)
                    except Exception:
                        pass
                    existing = None
                if not existing:
                    obj = CatalogModel(name='default', payload=DEFAULT_CATALOG, created_at=datetime.utcnow())
                    db.add(obj)
                    db.commit()
                # Migrate any in-memory catalogs into db if present
                for n, c in list(CATALOG_STORE.items()):
                    try:
                        exist = db.query(CatalogModel).filter(CatalogModel.name == n).first()
                        if not exist:
                            db.add(CatalogModel(name=n, payload=c, created_at=datetime.utcnow()))
                    except Exception:
                        db.rollback()
                db.commit()
                # Migrate in-memory projects if present
                for pid, pdata in list(INMEM_PROJECTS.items()):
                    try:
                        exist = db.query(ProjectModel).filter(ProjectModel.project_id == pid).first()
                        if not exist:
                            pm = ProjectModel(
                                project_id=pdata.get('project_id'),
                                name=pdata.get('name'),
                                plant_type=pdata.get('plant_type'),
                                standard=pdata.get('standard'),
                                voltage_levels=pdata.get('voltage_levels'),
                                service_condition=pdata.get('service_condition'),
                                created_at=datetime.utcnow()
                            )
                            db.add(pm)
                    except Exception:
                        db.rollback()
                db.commit()
                # Migrate in-memory cable results
                for pid, results in list(INMEM_CABLE_RESULTS.items()):
                    for rid, payload in results.items():
                        try:
                            exist = db.query(CableResultModel).filter(CableResultModel.result_id == rid, CableResultModel.project_id == pid).first()
                            if not exist:
                                obj = CableResultModel(
                                    result_id=rid,
                                    project_id=pid,
                                    cable_number=payload.get('cable_number', rid),
                                    payload=payload,
                                    created_at=datetime.utcnow()
                                )
                                db.add(obj)
                        except Exception:
                            db.rollback()
                db.commit()
            finally:
                db.close()
        else:
            if 'default' not in CATALOG_STORE:
                CATALOG_STORE['default'] = DEFAULT_CATALOG
    except Exception:
        pass

# ==================== API Endpoints ====================

@app.get("/")
async def root():
    return {"message": "SCEAP API Running", "version": "1.0.0", "db_usable": DB_USABLE}

@app.post("/api/v1/cable/single")
async def calculate_single_cable(cable: CableInput, catalog_name: Optional[str] = None, standard: Optional[str] = 'IEC') -> CableResult:
    """Calculate single cable"""
    try:
        # Use phase-based FLC and refine derating factors
        phase = (cable.phase_type or ('single' if cable.cores == 1 else 'three'))
        flc = calculate_flc_for_phase(cable.load_kw, cable.voltage, cable.pf, cable.efficiency, phase)
        grouping = derive_grouping_factor(cable.runs or 1, cable.feeder_type, phase)
        temp_factor = derive_temp_factor(cable.ambient_temp)
        installation_factor = 1.0
        # Apply standard-specific adjustments
        if standard and standard.lower().startswith('is'):
            # IS often applies slightly stricter derating defaults
            grouping = max(0.55, grouping * 0.95)
            temp_factor = round(max(0.85, temp_factor * 0.95), 3)
        derated = apply_derating(flc, grouping_factor=grouping, temp_factor=temp_factor, installation_factor=installation_factor)
        selected_size_str, size_value, ampacity, resistance_per_m, recommended_runs, recommended_cores = select_cable_size(derated, standard=standard, catalog_name=catalog_name, grouping=grouping, temp_factor=temp_factor)
        # Use recommended runs for voltage drop calculation if present, else the user-specified runs
        use_runs = recommended_runs if recommended_runs and recommended_runs > 0 else (cable.runs or 1)
        vdrop = calculate_voltage_drop(flc, cable.length, cable.voltage, resistance=resistance_per_m, size_mm2=size_value, runs=use_runs)
        od = calculate_cable_od(recommended_cores or 3, size_value)
        
        # Voltage drop limits per standard
        vd_limit = 5.0
        if standard and standard.lower().startswith('is'):
            # IS 1554 often uses stricter limits for underground/overhead; use 3% for high-voltage
            vd_limit = 3.0 if (cable.voltage and cable.voltage > 1000) else 5.0
        else:
            if cable.voltage and cable.voltage > 1000:
                vd_limit = 3.0

        vd_pass = (vdrop <= vd_limit)
        # Short-circuit check: if user supplied prospective_sc, verify adiabatic capacity
        sc_pass = check_short_circuit(size_value)
        if cable.prospective_sc:
            adi = adiabatic_short_circuit_capacity(size_value)
            sc_pass = adi >= float(cable.prospective_sc)
        ampacity_margin = None
        ampacity_margin_pct = None
        ampacity_base = None
        ampacity_corrected = None
        # if catalog_name provided and size matches an entry, compute base/corrected ampacity
        if catalog_name:
            # try to find matching catalog entry by size
            catalog_list = None
            if DB_USABLE:
                db = SessionLocal()
                try:
                    try:
                        row = db.query(CatalogModel).filter(CatalogModel.name == catalog_name).first()
                        if row and row.payload:
                            catalog_list = row.payload
                    except OperationalError:
                        catalog_list = CATALOG_STORE.get(catalog_name)
                finally:
                    db.close()
            else:
                catalog_list = CATALOG_STORE.get(catalog_name)

            if catalog_list:
                for entry in catalog_list:
                    if entry.get('size_mm2') and float(entry.get('size_mm2')) == float(size_value):
                        base_amp, corrected_amp, _ = compute_ampacity_from_entry(entry, standard=standard, grouping_factor=grouping, temp_factor=temp_factor)
                        ampacity_base = base_amp
                        ampacity_corrected = corrected_amp
                        break
        # fallback: use ampacity returned by select_cable_size
        if ampacity is not None and ampacity_base is None:
            try:
                ampacity_val = float(ampacity)
                ampacity_base = ampacity_val
                ampacity_corrected = ampacity_val
            except Exception:
                ampacity_val = None

        if ampacity_corrected is not None:
            try:
                ampacity_margin = round(float(ampacity_corrected) - derated, 2)
                ampacity_margin_pct = round(((float(ampacity_corrected) - derated) / float(ampacity_corrected)) * 100, 2) if float(ampacity_corrected) != 0 else None
            except Exception:
                pass

        ao = False
        if ampacity is not None and ampacity_margin is not None:
            ao = (ampacity_margin >= 0) and vd_pass and sc_pass
        # Build formulas metadata
        formulas = {
            'flc': 'I = P / (√3 × V × PF × η)',
            'derated': 'I_d = I / (grouping_factor × temp_factor × installation_factor)',
            'runs': 'runs = ceil(I_d / ampacity_corrected_of_one_conductor)',
            'vd': 'Vd% = (√3 × I_per_run × L × √(R^2 + X^2)) / V × 100',
            'adiabatic': 'I_adiabatic = K × sqrt(S / t)',
            'ampacity_correction': 'I_corr = I_base × grouping_factor × temp_factor × installation_factor'
        }

        return CableResult(
            id=cable.cable_number,
            cable_number=cable.cable_number,
            description=cable.description,
            flc=round(flc, 2),
            derated_current=round(derated, 2),
            selected_size=size_value,
            voltage_drop=round(vdrop, 2),
            sc_check=sc_pass,
            grouping_factor=grouping,
            status="pending",
            cores=recommended_cores or cable.cores or 3,
            od=od
            ,
            ampacity=ampacity if ampacity is not None else None,
            ampacity_base=ampacity_base,
            ampacity_corrected=ampacity_corrected,
            ampacity_margin=ampacity_margin,
            ampacity_margin_pct=ampacity_margin_pct,
            vd_limit=vd_limit,
            vd_pass=vd_pass
            ,
            ao=ao
            ,
            prospective_sc=cable.prospective_sc,
            standard=standard
            ,
            standard_ref = ('IEC 60287' if (standard and standard.lower().startswith('iec')) else ('IS 1554' if (standard and standard.lower().startswith('is')) else standard)),
            recommended_cores=recommended_cores,
            recommended_runs=recommended_runs,
            resistance_per_m=resistance_per_m,
            reactance_per_m=None,
            configuration=selected_size_str,
            formulas=formulas
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/v1/cable/bulk")
async def calculate_bulk_cables(cables: List[CableInput], catalog_name: Optional[str] = None, standard: Optional[str] = 'IEC') -> List[CableResult]:
    """Calculate multiple cables"""
    results = []
    for cable in cables:
        result = await calculate_single_cable(cable, catalog_name=catalog_name, standard=standard)
        results.append(result)
    return results

@app.post("/api/v1/cable/bulk_excel")
async def upload_cable_excel(file: UploadFile = File(...), catalog_name: Optional[str] = None) -> ExcelUploadResponse:
    """Upload Excel and process"""
    try:
        content = await file.read()
        cables, errors = parse_excel_cables(content)
        inputs_serialized = [c.dict() for c in cables]
        
        results = []
        for cable in cables:
            try:
                result = await calculate_single_cable(cable, catalog_name=catalog_name)
                results.append(result)
            except Exception as e:
                errors.append(f"{cable.cable_number}: {str(e)}")
        
        return ExcelUploadResponse(
            success=len(errors) == 0,
            cables_imported=len(results),
            errors=errors,
            results=results,
            inputs=inputs_serialized
        )
    except Exception as e:
        return ExcelUploadResponse(success=False, cables_imported=0, errors=[str(e)])


@app.post('/api/v1/cable/save_bulk')
async def save_bulk_results(project_id: str, results: List[CableResult]):
    """Save bulk cable results associated with a project (DB persistence)."""
    saved = 0
    if DB_USABLE:
        db = SessionLocal()
        try:
            for r in results:
                obj = CableResultModel(
                    result_id=r.id,
                    project_id=project_id,
                    cable_number=r.cable_number,
                    payload=r.dict(),
                    created_at=datetime.utcnow()
                )
                db.add(obj)
                saved += 1
            db.commit()
        except Exception as e:
            db.rollback()
            # Fallback to in-memory store on failure
            for r in results:
                INMEM_CABLE_RESULTS.setdefault(project_id, {})[r.id] = r.dict()
                saved += 1
        finally:
            db.close()
    else:
        for r in results:
            INMEM_CABLE_RESULTS.setdefault(project_id, {})[r.id] = r.dict()
            saved += 1

    return { 'saved': saved }


@app.post('/api/v1/catalogs/upload')
async def upload_catalog(file: UploadFile = File(...), name: Optional[str] = None):
    """Upload a catalog (Excel) and store it by name."""
    try:
        content = await file.read()
        catalog, errors = parse_catalog_excel(content)
        if errors:
            return { 'success': False, 'errors': errors, 'template_download': '/api/v1/catalogs/template' }
        if not name:
            name = f"catalog-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        if DB_USABLE:
            db = SessionLocal()
            try:
                obj = CatalogModel(name=name, payload=catalog, created_at=datetime.utcnow())
                db.add(obj)
                db.commit()
                return { 'success': True, 'name': name }
            except OperationalError:
                # DB/table not writable - fallback to in-memory store
                CATALOG_STORE[name] = catalog
                return { 'success': True, 'name': name }
            except Exception as e:
                db.rollback()
                raise HTTPException(status_code=500, detail=str(e))
            finally:
                db.close()
        else:
            CATALOG_STORE[name] = catalog
            return { 'success': True, 'name': name }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/api/v1/catalogs')
async def list_catalogs():
    """Return list of available catalogs (names)."""
    names = []
    if DB_USABLE:
        db = SessionLocal()
        try:
            try:
                rows = db.query(CatalogModel).all()
                names = [r.name for r in rows]
            except OperationalError:
                # Table possibly missing; try to create tables then re-query
                try:
                    Base.metadata.create_all(bind=Engine)
                except Exception:
                    pass
                try:
                    rows = db.query(CatalogModel).all()
                    names = [r.name for r in rows]
                except Exception:
                    # fallback to in-memory store on repeated failure
                    names = list(CATALOG_STORE.keys())
        finally:
            db.close()
    else:
        names = list(CATALOG_STORE.keys())
    return { 'catalogs': names }


@app.get('/api/v1/catalogs/template')
async def download_catalog_template():
    """Return a small XLSX template for catalog upload with headers and example rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(['size_mm2', 'ampacity', 'cores', 'ampacity_air', 'ampacity_ground', 'resistance_ohm_per_km', 'reactance_ohm_per_km', 'cable_dia_mm', 'xlpe', 'pairs', 'paralleled_count', 'conductor_material', 'insulation', 'sheath', 'formation'])
    ws.append([10, 55, 3, 55, 60, 1.83, 0.08, 20, 'XLPE', None, None, 'Cu', 'XLPE', 'PVC', '3C'])
    ws.append([16, 70, 3, 70, 75, 1.15, 0.07, 22, 'XLPE', None, None, 'Cu', 'XLPE', 'PVC', '3C'])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="catalog-template.xlsx"'
    }
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@app.get('/api/v1/import/template')
async def download_import_template():
    """Return an XLSX template for feeder/load list import matching the expected columns."""
    wb = Workbook()
    ws = wb.active
    headers = [
        'cable_number', 'description', 'load_kw', 'load_kva', 'voltage', 'pf', 'efficiency',
        'length', 'cable_type', 'from_equipment', 'to_equipment', 'breaker_type',
        'feeder_type', 'quantity', 'voltage_variation', 'power_supply', 'installation',
        'prospective_sc', 'phase_type', 'ambient_temp'
    ]
    ws.append(headers)
    # sample row
    ws.append(['C-001', 'Sample Motor', 55, 60, 415, 0.9, 0.95, 120, 'C', 'E1', 'E2', 'MCC', 'FDR', 1, None, '3ph', '30'])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="import-template.xlsx"'
    }
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@app.get('/api/v1/catalogs/{name}')
async def get_catalog(name: str):
    """Return the catalog payload by name."""
    if DB_USABLE:
        db = SessionLocal()
        try:
            try:
                row = db.query(CatalogModel).filter(CatalogModel.name == name).first()
            except OperationalError:
                # Create missing tables and retry once
                try:
                    Base.metadata.create_all(bind=Engine)
                except Exception:
                    pass
                row = None

            if not row:
                # fallback to in-memory store
                if name in CATALOG_STORE:
                    return { 'name': name, 'catalog': CATALOG_STORE[name] }
                raise HTTPException(status_code=404, detail='Catalog not found')
            return { 'name': row.name, 'catalog': row.payload }
        finally:
            db.close()
    else:
        if name not in CATALOG_STORE:
            raise HTTPException(status_code=404, detail='Catalog not found')
        return { 'name': name, 'catalog': CATALOG_STORE[name] }


@app.get('/api/v1/project/{project_id}/cables')
async def get_project_cables(project_id: str):
    if DB_USABLE:
        db = SessionLocal()
        try:
            rows = db.query(CableResultModel).filter(CableResultModel.project_id == project_id).all()
            return [r.payload for r in rows]
        finally:
            db.close()
    else:
        # return in-memory results
        results = INMEM_CABLE_RESULTS.get(project_id, {})
        return list(results.values())


@app.post('/api/v1/project/{project_id}/cable/{cable_id}/approve')
async def approve_cable(project_id: str, cable_id: str, status: str = 'approved'):
    """Update cable approval status in DB."""
    if DB_USABLE:
        db = SessionLocal()
        try:
            row = db.query(CableResultModel).filter(
                CableResultModel.project_id == project_id,
                CableResultModel.result_id == cable_id
            ).first()
            if row and row.payload:
                row.payload['status'] = status
                db.commit()
        except Exception:
            db.rollback()
        finally:
            db.close()
    else:
        # in-memory fallback
        project_store = INMEM_CABLE_RESULTS.get(project_id, {})
        if cable_id in project_store:
            project_store[cable_id]['status'] = status
            INMEM_CABLE_RESULTS[project_id] = project_store

    return { 'status': status, 'cable_id': cable_id }


@app.put('/api/v1/project/{project_id}/cable/{cable_number}')
async def upsert_project_cable(project_id: str, cable_number: str, payload: CableResult):
    """Create or update a cable result associated with a project."""
    if DB_USABLE:
        db = SessionLocal()
        try:
            row = db.query(CableResultModel).filter(
                CableResultModel.project_id == project_id,
                CableResultModel.result_id == payload.id
            ).first()
            if row:
                row.payload = payload.dict()
                row.created_at = datetime.utcnow()
                db.commit()
                return { 'updated': payload.cable_number }
            else:
                obj = CableResultModel(
                    result_id=payload.id,
                    project_id=project_id,
                    cable_number=payload.cable_number,
                    payload=payload.dict(),
                    created_at=datetime.utcnow()
                )
                db.add(obj)
                db.commit()
                return { 'created': payload.cable_number }
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            db.close()
    else:
        # Save to in-memory store (upsert behavior)
        project_store = INMEM_CABLE_RESULTS.setdefault(project_id, {})
        existed = payload.id in project_store
        project_store[payload.id] = payload.dict()
        INMEM_CABLE_RESULTS[project_id] = project_store
        return { 'updated' if existed else 'created': payload.cable_number }


@app.post("/api/v1/routing/auto", response_model=RoutingResult)
async def auto_route_cable(request: Dict[str, Any]):
    """Auto-route cable. Accepts simple payloads: {source, target, cable_id?} or full RoutingRequest fields."""
    try:
        source = request.get('source') or request.get('from_equipment') or request.get('from')
        target = request.get('target') or request.get('to_equipment') or request.get('to')
        cable_id = request.get('cable_id') or request.get('id') or 'CABLE-NA'

        path, length = routing_engine.find_shortest_path(source, target)
        fill_status = {}
        warnings = []

        for node in path:
            if node in routing_engine.tray_data:
                fill = routing_engine.tray_data[node]["fill"]
                fill_status[node] = fill
                if fill > 80:
                    warnings.append(f"{node} is {fill}% full")

        return RoutingResult(
            cable_id=cable_id,
            path=path,
            total_length=length,
            fill_status=fill_status,
            warnings=warnings
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/v1/routing/optimize", response_model=RoutingResult)
async def optimize_route(request: Dict[str, Any]):
    """Optimize route. Accepts {source, target, algorithm} where algorithm can be 'shortest' or 'least-fill'."""
    try:
        source = request.get('source') or request.get('from_equipment') or request.get('from')
        target = request.get('target') or request.get('to_equipment') or request.get('to')
        cable_id = request.get('cable_id') or 'CABLE-NA'
        algorithm = request.get('algorithm', 'shortest')

        if algorithm == 'least-fill':
            path, length = routing_engine.find_least_fill_path(source, target)
        else:
            path, length = routing_engine.find_shortest_path(source, target)

        fill_status = {}
        warnings = []
        for node in path:
            if node in routing_engine.tray_data:
                fill = routing_engine.tray_data[node]['fill']
                fill_status[node] = fill
                if fill > 80:
                    warnings.append(f"{node} is {fill}% full")

        return RoutingResult(
            cable_id=cable_id,
            path=path,
            total_length=length,
            fill_status=fill_status,
            warnings=warnings
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/v1/routing/trays")
async def get_tray_network():
    """Get tray network"""
    trays_list = []
    for k, v in routing_engine.tray_data.items():
        trays_list.append({ 'name': k, 'fill_percentage': v.get('fill', 0), 'capacity': v.get('capacity', 0) })

    avg = 0
    if trays_list:
        avg = sum(t['fill_percentage'] for t in trays_list) / len(trays_list)

    return {
        "trays": trays_list,
        "total_trays": len(trays_list),
        "average_fill": avg
    }


@app.get('/api/v1/routing/graph')
async def get_routing_graph():
    """Return graph nodes and edges for frontend visualization"""
    if not routing_engine.graph:
        return { 'nodes': [], 'edges': [] }

    nodes = []
    for n in routing_engine.graph.nodes():
        meta = routing_engine.tray_data.get(n, {})
        nodes.append({
            'id': n,
            'label': n,
            'meta': meta
        })

    edges = []
    for u, v, data in routing_engine.graph.edges(data=True):
        edges.append({ 'source': u, 'target': v, 'weight': float(data.get('weight', 1.0)) })

    return { 'nodes': nodes, 'edges': edges }

@app.post("/api/v1/project/setup")
async def setup_project(project: ProjectSetup):
    """Setup project (persist to DB when available)"""
    pid = f"PROJ-{int(datetime.utcnow().timestamp())}"
    created = datetime.utcnow()
    if DB_USABLE:
        db = SessionLocal()
        try:
            pm = ProjectModel(
                project_id=pid,
                name=project.project_name,
                plant_type=project.plant_type,
                standard=project.standard,
                voltage_levels=project.voltage_levels,
                service_condition=project.service_condition,
                created_at=created
            )
            db.add(pm)
            db.commit()
        except Exception:
            db.rollback()
        finally:
            db.close()
    else:
        INMEM_PROJECTS[pid] = {
            'project_id': pid,
            'name': project.project_name,
            'plant_type': project.plant_type,
            'standard': project.standard,
            'voltage_levels': project.voltage_levels,
            'service_condition': project.service_condition,
            'created_at': created.isoformat()
        }

    return {
        "project_id": pid,
        "created_at": created.isoformat(),
        "status": "initialized",
        "project": project
    }


@app.get('/api/v1/projects')
async def list_projects():
    if DB_USABLE:
        db = SessionLocal()
        try:
            rows = db.query(ProjectModel).all()
            return [{ 'project_id': r.project_id, 'name': r.name, 'plant_type': r.plant_type, 'standard': r.standard } for r in rows]
        finally:
            db.close()
    else:
        return list(INMEM_PROJECTS.values())

@app.get("/api/v1/standards")
async def get_standards():
    """Get standards"""
    return {
        "IEC": {"standard_number": "IEC 60287", "title": "Calculation of continuous current rating"},
        "IS": {"standard_number": "IS 1554", "title": "Cross-linked polyethylene insulated cables"}
    }

@app.get("/api/v1/cable-sizes")
async def get_cable_sizes(standard: str = "IEC"):
    """Get cable sizes"""
    return {
        "standard": standard,
        "sizes": [
            {"amps": 10, "size": "1.5"}, {"amps": 16, "size": "2.5"},
            {"amps": 25, "size": "4"}, {"amps": 35, "size": "6"},
            {"amps": 50, "size": "10"}, {"amps": 63, "size": "16"},
            {"amps": 80, "size": "25"}, {"amps": 100, "size": "35"},
            {"amps": 125, "size": "50"}, {"amps": 160, "size": "70"},
            {"amps": 200, "size": "95"}, {"amps": 250, "size": "120"},
            {"amps": 315, "size": "150"}, {"amps": 400, "size": "185"},
            {"amps": 500, "size": "240"}
        ]
    }

@app.get("/api/v1/health")
async def health_check():
    """Health check"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "features": {
            "excel_import": load_workbook is not None,
            "routing": nx is not None
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
